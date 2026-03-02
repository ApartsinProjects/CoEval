"""Complete Report (Excel) — REQ-A-7.1."""
from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from ..loader import EESDataModel
from ..metrics import (
    compute_all_agreements,
    compute_judge_scores,
    compute_teacher_scores,
    compute_student_scores,
    kappa_label,
)


def write_complete_report(model: EESDataModel, out_path: Path) -> Path:
    """Write the complete Excel workbook (REQ-A-7.1).

    Parameters
    ----------
    model:
        Loaded EES data model.
    out_path:
        Output file path (should end in .xlsx).

    Returns
    -------
    Path to the written workbook.
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side,
        )
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import ColorScaleRule
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for Excel reports. "
            "Install it with: pip install openpyxl"
        ) from exc

    wb = openpyxl.Workbook()
    units = model.units
    students = model.students
    teachers = model.teachers
    judges = model.judges

    # ------------------------------------------------------------------
    # Pre-compute metrics
    # ------------------------------------------------------------------
    agreements = compute_all_agreements(units, judges)
    judge_scores_map = compute_judge_scores(units, judges, agreements)
    teacher_scores_map = compute_teacher_scores(units, teachers, students)
    student_scores_map = compute_student_scores(units, students, model.datapoints)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    def _avg(lst):
        return sum(lst) / len(lst) if lst else None

    def _std(lst):
        if len(lst) < 2:
            return None
        mean = sum(lst) / len(lst)
        var = sum((x - mean) ** 2 for x in lst) / (len(lst) - 1)
        return var ** 0.5

    def _style_header(ws):
        """Freeze header row and apply header style."""
        ws.freeze_panes = 'A2'
        header_fill = PatternFill('solid', fgColor='DDE8F0')
        header_font = Font(bold=True, size=10)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='left')

    def _style_column_widths(ws):
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in col
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                max(8, max_len + 2), 40
            )

    def _alt_row(ws, n_rows):
        """Alternating row shading."""
        alt_fill = PatternFill('solid', fgColor='F5F7FA')
        for row_idx in range(2, n_rows + 2):  # rows are 1-indexed, skip header
            if row_idx % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = alt_fill

    def _score_to_float(v):
        if isinstance(v, (int, float)):
            return float(v)
        return None

    # ------------------------------------------------------------------
    # Sheet 1 — Raw Slice Summary
    # ------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = 'Raw Slice Summary'

    headers1 = [
        'Task ID', 'Teacher Model ID', 'Student Model ID', 'Judge Model ID',
        'Rubric Aspect', 'Total Data Points', 'Valid Data Points',
        'Invalid Data Points', 'Average Score', 'Score Std Dev',
        'High Count', 'Medium Count', 'Low Count',
        'Self Judging', 'Self Teaching',
    ]
    ws1.append(headers1)

    # Group units by (task, teacher, student, judge, aspect)
    slice_data: dict[tuple, list[float]] = defaultdict(list)
    slice_valid: dict[tuple, int] = defaultdict(int)
    slice_total: dict[tuple, int] = defaultdict(int)
    slice_hist: dict[tuple, dict] = defaultdict(lambda: {'High': 0, 'Medium': 0, 'Low': 0})
    slice_self: dict[tuple, dict] = defaultdict(lambda: {'sj': False, 'st': False})

    for u in units:
        key = (u.task_id, u.teacher_model_id, u.student_model_id,
               u.judge_model_id, u.rubric_aspect)
        slice_data[key].append(u.score_norm)
        slice_valid[key] += 1
        slice_hist[key][u.score] = slice_hist[key].get(u.score, 0) + 1
        if u.is_self_judging:
            slice_self[key]['sj'] = True
        if u.is_self_teaching:
            slice_self[key]['st'] = True

    # Total includes invalid records
    for rec in model.eval_records:
        for aspect in model.rubrics.get(rec.task_id, {}):
            key = (rec.task_id, rec.teacher_model_id, rec.student_model_id,
                   rec.judge_model_id, aspect)
            slice_total[key] += 1

    rows1 = []
    for key in sorted(slice_data.keys()):
        (task_id, teacher_id, student_id, judge_id, aspect) = key
        scores = slice_data[key]
        valid = slice_valid[key]
        total = max(slice_total.get(key, valid), valid)
        invalid = total - valid
        avg = _avg(scores)
        std = _std(scores)
        hist = slice_hist[key]
        sj = slice_self[key]['sj']
        st = slice_self[key]['st']
        rows1.append([
            task_id, teacher_id, student_id, judge_id, aspect,
            total, valid, invalid,
            round(avg, 4) if avg is not None else None,
            round(std, 4) if std is not None else None,
            hist.get('High', 0), hist.get('Medium', 0), hist.get('Low', 0),
            '⚠' if sj else '', '⚠' if st else '',
        ])
    for row in rows1:
        ws1.append(row)

    _style_header(ws1)
    _alt_row(ws1, len(rows1))
    _style_column_widths(ws1)

    # Conditional colour scale on Average Score (column 9)
    avg_col = 'I'
    if rows1:
        ws1.conditional_formatting.add(
            f'{avg_col}2:{avg_col}{len(rows1)+1}',
            ColorScaleRule(
                start_type='num', start_value=0, start_color='FF4444',
                mid_type='num', mid_value=0.5, mid_color='FFFF44',
                end_type='num', end_value=1, end_color='44AA44',
            ),
        )

    # ------------------------------------------------------------------
    # Sheet 2 — Aggregated by (Task, Teacher, Student, Judge)
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet('Aggregated by Model Combo')
    headers2 = [
        'Task ID', 'Teacher Model ID', 'Student Model ID', 'Judge Model ID',
        'Total Data Points', 'Valid Data Points', 'Invalid Data Points',
        'Average Score', 'Score Std Dev',
        'High Count', 'Medium Count', 'Low Count',
    ]
    ws2.append(headers2)

    combo_data: dict[tuple, list[float]] = defaultdict(list)
    combo_hist: dict[tuple, dict] = defaultdict(lambda: {'High': 0, 'Medium': 0, 'Low': 0})

    for u in units:
        key4 = (u.task_id, u.teacher_model_id, u.student_model_id, u.judge_model_id)
        combo_data[key4].append(u.score_norm)
        combo_hist[key4][u.score] = combo_hist[key4].get(u.score, 0) + 1

    rows2 = []
    for key4 in sorted(combo_data.keys()):
        task_id, teacher_id, student_id, judge_id = key4
        scores = combo_data[key4]
        valid = len(scores)
        total = valid  # best estimate
        hist = combo_hist[key4]
        avg = _avg(scores)
        std = _std(scores)
        rows2.append([
            task_id, teacher_id, student_id, judge_id,
            total, valid, 0,
            round(avg, 4) if avg is not None else None,
            round(std, 4) if std is not None else None,
            hist.get('High', 0), hist.get('Medium', 0), hist.get('Low', 0),
        ])
    for row in rows2:
        ws2.append(row)
    _style_header(ws2)
    _alt_row(ws2, len(rows2))
    _style_column_widths(ws2)

    if rows2:
        ws2.conditional_formatting.add(
            f'H2:H{len(rows2)+1}',
            ColorScaleRule(
                start_type='num', start_value=0, start_color='FF4444',
                mid_type='num', mid_value=0.5, mid_color='FFFF44',
                end_type='num', end_value=1, end_color='44AA44',
            ),
        )

    # ------------------------------------------------------------------
    # Sheet 3 — Model Summary
    # ------------------------------------------------------------------
    ws3 = wb.create_sheet('Model Summary')
    headers3 = [
        'Model ID', 'Roles',
        'As Student — Avg Score', 'As Student — Valid Evals',
        'As Teacher — Score (V1)', 'As Teacher — Score (S2)', 'As Teacher — Score (R3)',
        'As Judge — SPA', 'As Judge — WPA', 'As Judge — Kappa',
    ]
    ws3.append(headers3)

    all_models = sorted(set(teachers + students + judges))
    config_models = {m.get('name', ''): m for m in model.config.get('models', [])}

    rows3 = []
    for mid in all_models:
        roles_list = []
        cmodel = config_models.get(mid, {})
        for r in cmodel.get('roles', []):
            if r not in roles_list:
                roles_list.append(r)
        # Infer roles from dimensions
        if mid in teachers and 'teacher' not in roles_list:
            roles_list.append('teacher')
        if mid in students and 'student' not in roles_list:
            roles_list.append('student')
        if mid in judges and 'judge' not in roles_list:
            roles_list.append('judge')

        # Student metrics
        sr = student_scores_map.get(mid)
        student_avg = round(sr.overall, 4) if sr and sr.overall is not None else None
        student_evals = sr.valid_evals if sr else 0

        # Teacher metrics
        tr = teacher_scores_map.get(mid)
        t_v1 = round(tr.v1, 4) if tr else None
        t_s2 = round(tr.s2, 4) if tr else None
        t_r3 = round(tr.r3, 4) if tr else None

        # Judge metrics
        jr = judge_scores_map.get(mid)
        j_spa = round(jr.spa_mean, 4) if jr and jr.spa_mean is not None else None
        j_wpa = round(jr.wpa_mean, 4) if jr and jr.wpa_mean is not None else None
        j_k = round(jr.kappa_mean, 4) if jr and jr.kappa_mean is not None else None

        rows3.append([
            mid, ', '.join(roles_list),
            student_avg, student_evals,
            t_v1, t_s2, t_r3,
            j_spa, j_wpa, j_k,
        ])

    for row in rows3:
        ws3.append(row)
    _style_header(ws3)
    _alt_row(ws3, len(rows3))
    _style_column_widths(ws3)

    # ------------------------------------------------------------------
    # Save
    # ------------------------------------------------------------------
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path
