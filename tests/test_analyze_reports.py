"""Smoke tests for coeval.analyze.reports and coeval.analyze.main.

For each report writer:
  - verifies the expected output file(s) are created
  - verifies output has non-trivial content
  - verifies key structural markers are present

Plotly.js downloads are avoided by creating a fake plotly.min.js in the
COEVAL_CACHE directory.  All callers of get_plotly_js() (regardless of which
module they're imported into) find the cached file automatically.
"""
from __future__ import annotations

import json
import os

import pytest
import yaml

from coeval.analyze.loader import AnalyticalUnit, EESDataModel, SCORE_MAP


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def fake_plotly_cache(tmp_path, monkeypatch):
    """
    Point COEVAL_CACHE to a tmp directory with a pre-populated plotly.min.js.
    This prevents get_plotly_js() from making any network calls.
    """
    cache_base = tmp_path / 'coeval_cache'
    (cache_base / 'plotly').mkdir(parents=True)
    plotly_file = cache_base / 'plotly' / 'plotly.min.js'
    plotly_file.write_text('/* fake plotly stub */\nPlotly = {};', encoding='utf-8')
    monkeypatch.setenv('COEVAL_CACHE', str(cache_base))
    return plotly_file


def _unit(
    response_id='r1',
    datapoint_id='dp1',
    task_id='task1',
    teacher='T1',
    student='S1',
    judge='J1',
    aspect='acc',
    score='High',
) -> AnalyticalUnit:
    return AnalyticalUnit(
        response_id=response_id,
        datapoint_id=datapoint_id,
        task_id=task_id,
        teacher_model_id=teacher,
        student_model_id=student,
        judge_model_id=judge,
        rubric_aspect=aspect,
        score=score,
        score_norm=SCORE_MAP.get(score, 0.0),
        is_self_judging=(judge == student),
        is_self_teaching=(teacher == student),
        evaluated_at='2024-01-01T00:00:00Z',
    )


@pytest.fixture
def minimal_model(tmp_path) -> EESDataModel:
    """
    Two judges (J1, J2), two students (S1, S2), one teacher (T1),
    two datapoints (dp1, dp2), two aspects (acc, fmt).
    Provides enough dimensionality for all report types.
    """
    units = []
    scores_matrix = {
        ('S1', 'J1'): 'High',
        ('S1', 'J2'): 'High',
        ('S2', 'J1'): 'Low',
        ('S2', 'J2'): 'Medium',
    }
    for dp_id in ('dp1', 'dp2'):
        for (student, judge), score in scores_matrix.items():
            for aspect in ('acc', 'fmt'):
                units.append(_unit(
                    response_id=f'{dp_id}-{student}-{judge}-{aspect}',
                    datapoint_id=dp_id,
                    teacher='T1',
                    student=student,
                    judge=judge,
                    aspect=aspect,
                    score=score,
                ))

    datapoints = {
        'dp1': {
            'id': 'dp1', 'teacher_model_id': 'T1', 'task_id': 'task1',
            'sampled_target_attributes': {'sentiment': 'positive'},
        },
        'dp2': {
            'id': 'dp2', 'teacher_model_id': 'T1', 'task_id': 'task1',
            'sampled_target_attributes': {'sentiment': 'negative'},
        },
    }

    return EESDataModel(
        run_path=tmp_path,
        meta={'experiment_id': 'test-exp', 'status': 'completed'},
        config={
            'models': [
                {'name': 'T1', 'roles': ['teacher']},
                {'name': 'S1', 'roles': ['student']},
                {'name': 'S2', 'roles': ['student']},
                {'name': 'J1', 'roles': ['judge']},
                {'name': 'J2', 'roles': ['judge']},
            ],
            'experiment': {'id': 'test-exp'},
        },
        rubrics={'task1': {'acc': 'Accuracy', 'fmt': 'Format'}},
        datapoints=datapoints,
        responses={},
        eval_records=[],
        units=units,
        tasks=['task1'],
        teachers=['T1'],
        students=['S1', 'S2'],
        judges=['J1', 'J2'],
        aspects_by_task={'task1': ['acc', 'fmt']},
        target_attrs_by_task={'task1': {'sentiment': ['negative', 'positive']}},
        total_records=len(units),
        valid_records=len(units),
        self_judging_count=0,
        self_teaching_count=0,
        both_count=0,
        load_warnings=[],
        is_partial=False,
    )


# ---------------------------------------------------------------------------
# Shared assertions
# ---------------------------------------------------------------------------

def _assert_html(path, *expected_fragments):
    """Assert index.html exists, has meaningful content, and contains key markers."""
    assert path.exists(), f"Expected {path} to exist"
    text = path.read_text(encoding='utf-8')
    assert len(text) > 300, f"HTML suspiciously short ({len(text)} chars): {path}"
    for frag in expected_fragments:
        assert frag in text, f"Expected '{frag}' in {path.name}"


def _assert_html_has_data_const(path):
    text = path.read_text(encoding='utf-8')
    assert 'const DATA' in text or 'window.DATA' in text or 'var DATA' in text, \
        f"No DATA constant found in {path}"


# ---------------------------------------------------------------------------
# Complete Report (Excel) — REQ-A-7.1
# ---------------------------------------------------------------------------

class TestCompleteReport:

    def test_creates_xlsx_file(self, tmp_path, minimal_model):
        from coeval.analyze.reports.excel import write_complete_report
        out = tmp_path / 'report.xlsx'
        write_complete_report(minimal_model, out)
        assert out.exists()
        assert out.stat().st_size > 1000

    def test_xlsx_has_multiple_sheets(self, tmp_path, minimal_model):
        import openpyxl
        from coeval.analyze.reports.excel import write_complete_report
        out = tmp_path / 'report.xlsx'
        write_complete_report(minimal_model, out)
        wb = openpyxl.load_workbook(out)
        assert len(wb.sheetnames) >= 2

    def test_first_sheet_has_data_rows(self, tmp_path, minimal_model):
        import openpyxl
        from coeval.analyze.reports.excel import write_complete_report
        out = tmp_path / 'report.xlsx'
        write_complete_report(minimal_model, out)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.max_row > 1      # header + at least one data row

    def test_empty_units_still_creates_file(self, tmp_path, minimal_model):
        import openpyxl
        from coeval.analyze.reports.excel import write_complete_report
        minimal_model.units.clear()
        out = tmp_path / 'report_empty.xlsx'
        write_complete_report(minimal_model, out)
        assert out.exists()
        wb = openpyxl.load_workbook(out)
        assert len(wb.sheetnames) >= 1


# ---------------------------------------------------------------------------
# Coverage Summary — REQ-A-7.9
# ---------------------------------------------------------------------------

class TestCoverageSummary:

    def test_creates_index_html(self, tmp_path, minimal_model, fake_plotly_cache):
        from coeval.analyze.reports.coverage import write_coverage_summary
        out_dir = tmp_path / 'coverage'
        write_coverage_summary(minimal_model, out_dir)
        _assert_html(out_dir / 'index.html', 'Coverage Summary', 'DATA')

    def test_html_has_plotly_script_tag(self, tmp_path, minimal_model, fake_plotly_cache):
        from coeval.analyze.reports.coverage import write_coverage_summary
        out_dir = tmp_path / 'coverage'
        write_coverage_summary(minimal_model, out_dir)
        html = (out_dir / 'index.html').read_text(encoding='utf-8')
        assert 'plotly' in html.lower()

    def test_shared_plotly_parameter_copies_file(self, tmp_path, minimal_model, fake_plotly_cache):
        """shared_plotly bypasses get_plotly_js(); file is copied from shared."""
        from coeval.analyze.reports.coverage import write_coverage_summary
        out_dir = tmp_path / 'coverage'
        write_coverage_summary(minimal_model, out_dir, shared_plotly=fake_plotly_cache)
        # plotly.min.js should have been copied into out_dir
        assert (out_dir / 'plotly.min.js').exists()

    def test_creates_output_directory(self, tmp_path, minimal_model, fake_plotly_cache):
        from coeval.analyze.reports.coverage import write_coverage_summary
        out_dir = tmp_path / 'new_dir' / 'coverage'
        write_coverage_summary(minimal_model, out_dir)
        assert out_dir.is_dir()


# ---------------------------------------------------------------------------
# Score Distribution — REQ-A-7.3
# ---------------------------------------------------------------------------

class TestScoreDistribution:

    def test_creates_index_html(self, tmp_path, minimal_model, fake_plotly_cache):
        from coeval.analyze.reports.score_dist import write_score_distribution
        out_dir = tmp_path / 'score_dist'
        write_score_distribution(minimal_model, out_dir)
        _assert_html(out_dir / 'index.html', 'Score Distribution', 'DATA')

    def test_html_has_data_constant(self, tmp_path, minimal_model, fake_plotly_cache):
        from coeval.analyze.reports.score_dist import write_score_distribution
        out_dir = tmp_path / 'score_dist'
        write_score_distribution(minimal_model, out_dir)
        _assert_html_has_data_const(out_dir / 'index.html')


# ---------------------------------------------------------------------------
# Judge Report — REQ-A-7.5
# ---------------------------------------------------------------------------

class TestJudgeReport:

    def test_creates_index_html(self, tmp_path, minimal_model, fake_plotly_cache):
        from coeval.analyze.reports.judge_report import write_judge_report
        out_dir = tmp_path / 'judge_report'
        write_judge_report(minimal_model, out_dir)
        _assert_html(out_dir / 'index.html', 'DATA')

    def test_single_judge_still_creates_html(self, tmp_path, minimal_model, fake_plotly_cache):
        """With 1 judge the degenerate notice path is exercised."""
        from coeval.analyze.reports.judge_report import write_judge_report
        # Narrow to one judge
        minimal_model.judges[:] = ['J1']
        minimal_model.units[:] = [u for u in minimal_model.units
                                   if u.judge_model_id == 'J1']
        out_dir = tmp_path / 'judge_single'
        write_judge_report(minimal_model, out_dir)
        assert (out_dir / 'index.html').exists()
        html = (out_dir / 'index.html').read_text(encoding='utf-8')
        assert len(html) > 100

    def test_zero_judges_still_creates_html(self, tmp_path, minimal_model, fake_plotly_cache):
        from coeval.analyze.reports.judge_report import write_judge_report
        minimal_model.judges[:] = []
        minimal_model.units[:] = []
        out_dir = tmp_path / 'judge_zero'
        write_judge_report(minimal_model, out_dir)
        assert (out_dir / 'index.html').exists()


# ---------------------------------------------------------------------------
# Teacher Report — REQ-A-7.4
# ---------------------------------------------------------------------------

class TestTeacherReport:

    def test_creates_index_html(self, tmp_path, minimal_model, fake_plotly_cache):
        from coeval.analyze.reports.teacher_report import write_teacher_report
        out_dir = tmp_path / 'teacher_report'
        write_teacher_report(minimal_model, out_dir)
        _assert_html(out_dir / 'index.html', 'DATA')

    def test_single_student_exercises_degenerate_path(self, tmp_path, minimal_model,
                                                        fake_plotly_cache):
        from coeval.analyze.reports.teacher_report import write_teacher_report
        minimal_model.students[:] = ['S1']
        minimal_model.units[:] = [u for u in minimal_model.units
                                   if u.student_model_id == 'S1']
        out_dir = tmp_path / 'teacher_single_student'
        write_teacher_report(minimal_model, out_dir)
        assert (out_dir / 'index.html').exists()


# ---------------------------------------------------------------------------
# Student Report — REQ-A-7.6
# ---------------------------------------------------------------------------

class TestStudentReport:

    def test_creates_index_html(self, tmp_path, minimal_model, fake_plotly_cache):
        from coeval.analyze.reports.student_report import write_student_report
        out_dir = tmp_path / 'student_report'
        write_student_report(minimal_model, out_dir)
        _assert_html(out_dir / 'index.html', 'DATA')

    def test_html_has_data_constant(self, tmp_path, minimal_model, fake_plotly_cache):
        from coeval.analyze.reports.student_report import write_student_report
        out_dir = tmp_path / 'student_report'
        write_student_report(minimal_model, out_dir)
        _assert_html_has_data_const(out_dir / 'index.html')


# ---------------------------------------------------------------------------
# Interaction Matrix — REQ-A-7.7
# ---------------------------------------------------------------------------

class TestInteractionMatrix:

    def test_creates_index_html(self, tmp_path, minimal_model, fake_plotly_cache):
        from coeval.analyze.reports.interaction import write_interaction_matrix
        out_dir = tmp_path / 'interaction'
        write_interaction_matrix(minimal_model, out_dir)
        _assert_html(out_dir / 'index.html', 'DATA')


# ---------------------------------------------------------------------------
# Judge Consistency — REQ-A-7.8
# ---------------------------------------------------------------------------

class TestJudgeConsistency:

    def test_creates_index_html(self, tmp_path, minimal_model, fake_plotly_cache):
        from coeval.analyze.reports.consistency import write_judge_consistency
        out_dir = tmp_path / 'consistency'
        write_judge_consistency(minimal_model, out_dir)
        _assert_html(out_dir / 'index.html', 'DATA')

    def test_html_is_well_formed(self, tmp_path, minimal_model, fake_plotly_cache):
        from coeval.analyze.reports.consistency import write_judge_consistency
        out_dir = tmp_path / 'consistency'
        write_judge_consistency(minimal_model, out_dir)
        html = (out_dir / 'index.html').read_text(encoding='utf-8')
        assert html.startswith('<!DOCTYPE html>')
        assert '</html>' in html


# ---------------------------------------------------------------------------
# Robust Summary — REQ-A-7.10
# ---------------------------------------------------------------------------

class TestRobustSummary:

    def test_creates_index_html_low_threshold(self, tmp_path, minimal_model, fake_plotly_cache):
        """threshold=0.0 → D_robust is non-empty → full report generated."""
        from coeval.analyze.reports.robust import write_robust_summary
        out_dir = tmp_path / 'robust'
        write_robust_summary(
            minimal_model, out_dir,
            judge_selection='all',
            agreement_metric='spa',
            agreement_threshold=0.0,
            teacher_score_formula='v1',
        )
        _assert_html(out_dir / 'index.html', 'DATA')

    def test_empty_d_robust_exits(self, tmp_path, minimal_model, fake_plotly_cache):
        """Impossible threshold → D_robust empty → should sys.exit(0) or handle gracefully."""
        from coeval.analyze.reports.robust import write_robust_summary
        # Make all units conflicting so consistency fraction is always 0
        for u in minimal_model.units:
            object.__setattr__(u, 'score',
                               'High' if u.judge_model_id == 'J1' else 'Low')
            object.__setattr__(u, 'score_norm',
                               1.0 if u.judge_model_id == 'J1' else 0.0)
        out_dir = tmp_path / 'robust_empty'
        # threshold=1.0 with conflicting judges → D_robust = ∅ → should exit cleanly
        with pytest.raises(SystemExit):
            write_robust_summary(
                minimal_model, out_dir,
                judge_selection='all',
                agreement_metric='spa',
                agreement_threshold=1.0,
                teacher_score_formula='v1',
            )


# ---------------------------------------------------------------------------
# Export Benchmark — REQ-A-7.11
# ---------------------------------------------------------------------------

class TestExportBenchmark:

    def test_creates_jsonl_file(self, tmp_path, minimal_model):
        from coeval.analyze.reports.export_benchmark import export_benchmark
        out_path = tmp_path / 'benchmark.jsonl'
        export_benchmark(
            minimal_model, out_path,
            judge_selection='all',
            agreement_metric='spa',
            agreement_threshold=0.0,
            teacher_score_formula='v1',
            benchmark_format='jsonl',
        )
        assert out_path.exists()

    def test_jsonl_has_schema_version(self, tmp_path, minimal_model):
        from coeval.analyze.reports.export_benchmark import export_benchmark
        out_path = tmp_path / 'benchmark.jsonl'
        export_benchmark(
            minimal_model, out_path,
            judge_selection='all',
            agreement_metric='spa',
            agreement_threshold=0.0,
            teacher_score_formula='v1',
            benchmark_format='jsonl',
        )
        if out_path.exists() and out_path.stat().st_size > 0:
            first_line = out_path.read_text(encoding='utf-8').splitlines()[0]
            record = json.loads(first_line)
            assert record.get('schema_version') == 'coeval-benchmark-v1'

    def test_jsonl_records_have_required_fields(self, tmp_path, minimal_model):
        from coeval.analyze.reports.export_benchmark import export_benchmark
        out_path = tmp_path / 'benchmark.jsonl'
        export_benchmark(
            minimal_model, out_path,
            judge_selection='all',
            agreement_metric='spa',
            agreement_threshold=0.0,
            teacher_score_formula='v1',
            benchmark_format='jsonl',
        )
        if out_path.exists() and out_path.stat().st_size > 0:
            for line in out_path.read_text(encoding='utf-8').splitlines():
                if not line.strip():
                    continue
                rec = json.loads(line)
                assert 'schema_version' in rec
                assert 'datapoint_id' in rec

    def test_empty_d_robust_raises_system_exit(self, tmp_path, minimal_model):
        """If D_robust = ∅, export_benchmark raises SystemExit(1) (REQ-A-5.7.2a)."""
        from coeval.analyze.reports.export_benchmark import export_benchmark
        # Make all units conflicting (J1=High vs J2=Low) to force D_robust = ∅
        for u in minimal_model.units:
            object.__setattr__(u, 'score',
                               'High' if u.judge_model_id == 'J1' else 'Low')
            object.__setattr__(u, 'score_norm',
                               1.0 if u.judge_model_id == 'J1' else 0.0)
        out_path = tmp_path / 'empty.jsonl'
        with pytest.raises(SystemExit) as exc_info:
            export_benchmark(
                minimal_model, out_path,
                judge_selection='all',
                agreement_metric='spa',
                agreement_threshold=1.0,
                teacher_score_formula='v1',
                benchmark_format='jsonl',
            )
        assert exc_info.value.code == 1

    @pytest.mark.skipif(
        not __import__('importlib').util.find_spec('pyarrow'),
        reason='pyarrow not installed',
    )
    def test_creates_parquet_file(self, tmp_path, minimal_model):
        from coeval.analyze.reports.export_benchmark import export_benchmark
        out_path = tmp_path / 'benchmark.parquet'
        export_benchmark(
            minimal_model, out_path,
            judge_selection='all',
            agreement_metric='spa',
            agreement_threshold=0.0,
            teacher_score_formula='v1',
            benchmark_format='parquet',
        )
        assert out_path.exists()


# ---------------------------------------------------------------------------
# run_analyze() integration — REQ-A-8.1
# ---------------------------------------------------------------------------

def _make_ees_folder(tmp_path):
    """Build a minimal real EES folder on disk."""
    run = tmp_path / 'exp01'
    run.mkdir(exist_ok=True)

    (run / 'meta.json').write_text(
        json.dumps({'experiment_id': 'exp01', 'status': 'completed'}),
        encoding='utf-8',
    )

    cfg = {
        'models': [
            {'name': 'T1', 'interface': 'openai', 'parameters': {}, 'roles': ['teacher']},
            {'name': 'S1', 'interface': 'openai', 'parameters': {}, 'roles': ['student']},
            {'name': 'J1', 'interface': 'openai', 'parameters': {}, 'roles': ['judge']},
        ],
        'tasks': [{'name': 'task1', 'description': 'T', 'output_description': 'OD'}],
        'experiment': {'id': 'exp01', 'storage_folder': str(tmp_path)},
    }
    (run / 'config.yaml').write_text(yaml.dump(cfg), encoding='utf-8')

    (run / 'phase2_rubric').mkdir()
    (run / 'phase2_rubric' / 'task1.rubric.json').write_text(
        json.dumps({'accuracy': 'Is accurate?'}), encoding='utf-8',
    )

    (run / 'phase3_datapoints').mkdir()
    dps = [
        {
            'id': 'dp1', 'task_id': 'task1', 'teacher_model_id': 'T1',
            'prompt': 'P', 'response': 'R',
            'sampled_target_attributes': {'sentiment': 'pos'},
        },
    ]
    (run / 'phase3_datapoints' / 'T1.task1.datapoints.jsonl').write_text(
        '\n'.join(json.dumps(d) for d in dps), encoding='utf-8',
    )

    (run / 'phase4_responses').mkdir()
    resps = [
        {
            'id': 'resp1', 'datapoint_id': 'dp1', 'task_id': 'task1',
            'teacher_model_id': 'T1', 'student_model_id': 'S1',
            'response': 'A',
        },
    ]
    (run / 'phase4_responses' / 'S1.T1.task1.responses.jsonl').write_text(
        '\n'.join(json.dumps(r) for r in resps), encoding='utf-8',
    )

    (run / 'phase5_evaluations').mkdir()
    evals = [
        {
            'id': 'ev1', 'response_id': 'resp1', 'datapoint_id': 'dp1',
            'task_id': 'task1', 'teacher_model_id': 'T1', 'judge_model_id': 'J1',
            'scores': {'accuracy': 'High'}, 'evaluated_at': '2024-01-01T00:00:00Z',
        },
    ]
    (run / 'phase5_evaluations' / 'J1.T1.task1.evaluations.jsonl').write_text(
        '\n'.join(json.dumps(e) for e in evals), encoding='utf-8',
    )
    return run


class TestRunAnalyze:

    def test_missing_run_path_returns_1(self, tmp_path, fake_plotly_cache):
        from coeval.analyze.main import run_analyze
        code = run_analyze(
            run_path=str(tmp_path / 'does_not_exist'),
            out_path=str(tmp_path / 'out'),
            subcommand='coverage-summary',
        )
        assert code == 1

    def test_complete_report_returns_0_and_creates_xlsx(self, tmp_path, fake_plotly_cache):
        from coeval.analyze.main import run_analyze
        run = _make_ees_folder(tmp_path)
        out = tmp_path / 'out.xlsx'
        code = run_analyze(
            run_path=str(run),
            out_path=str(out),
            subcommand='complete-report',
            partial_ok=True,
        )
        assert code == 0
        assert out.exists()

    def test_coverage_summary_returns_0_and_creates_html(self, tmp_path, fake_plotly_cache):
        from coeval.analyze.main import run_analyze
        run = _make_ees_folder(tmp_path)
        out_dir = tmp_path / 'coverage_out'
        code = run_analyze(
            run_path=str(run),
            out_path=str(out_dir),
            subcommand='coverage-summary',
            partial_ok=True,
        )
        assert code == 0
        assert (out_dir / 'index.html').exists()

    def test_score_distribution_returns_0(self, tmp_path, fake_plotly_cache):
        from coeval.analyze.main import run_analyze
        run = _make_ees_folder(tmp_path)
        out_dir = tmp_path / 'sd_out'
        code = run_analyze(
            run_path=str(run),
            out_path=str(out_dir),
            subcommand='score-distribution',
            partial_ok=True,
        )
        assert code == 0

    def test_teacher_report_returns_0(self, tmp_path, fake_plotly_cache):
        from coeval.analyze.main import run_analyze
        run = _make_ees_folder(tmp_path)
        code = run_analyze(
            run_path=str(run),
            out_path=str(tmp_path / 'tr_out'),
            subcommand='teacher-report',
            partial_ok=True,
        )
        assert code == 0

    def test_judge_report_returns_0(self, tmp_path, fake_plotly_cache):
        from coeval.analyze.main import run_analyze
        run = _make_ees_folder(tmp_path)
        code = run_analyze(
            run_path=str(run),
            out_path=str(tmp_path / 'jr_out'),
            subcommand='judge-report',
            partial_ok=True,
        )
        assert code == 0

    def test_student_report_returns_0(self, tmp_path, fake_plotly_cache):
        from coeval.analyze.main import run_analyze
        run = _make_ees_folder(tmp_path)
        code = run_analyze(
            run_path=str(run),
            out_path=str(tmp_path / 'sr_out'),
            subcommand='student-report',
            partial_ok=True,
        )
        assert code == 0

    def test_interaction_matrix_returns_0(self, tmp_path, fake_plotly_cache):
        from coeval.analyze.main import run_analyze
        run = _make_ees_folder(tmp_path)
        code = run_analyze(
            run_path=str(run),
            out_path=str(tmp_path / 'im_out'),
            subcommand='interaction-matrix',
            partial_ok=True,
        )
        assert code == 0

    def test_judge_consistency_returns_0(self, tmp_path, fake_plotly_cache):
        from coeval.analyze.main import run_analyze
        run = _make_ees_folder(tmp_path)
        code = run_analyze(
            run_path=str(run),
            out_path=str(tmp_path / 'jc_out'),
            subcommand='judge-consistency',
            partial_ok=True,
        )
        assert code == 0

    def test_export_benchmark_returns_0(self, tmp_path, fake_plotly_cache):
        from coeval.analyze.main import run_analyze
        run = _make_ees_folder(tmp_path)
        out_path = tmp_path / 'benchmark.jsonl'
        code = run_analyze(
            run_path=str(run),
            out_path=str(out_path),
            subcommand='export-benchmark',
            agreement_threshold=0.0,
            partial_ok=True,
        )
        assert code == 0

    def test_unknown_subcommand_returns_1(self, tmp_path, fake_plotly_cache):
        from coeval.analyze.main import run_analyze
        run = _make_ees_folder(tmp_path)
        code = run_analyze(
            run_path=str(run),
            out_path=str(tmp_path / 'out'),
            subcommand='nonexistent-command',
        )
        assert code == 1

    def test_all_subcommand_creates_multiple_reports(self, tmp_path, fake_plotly_cache):
        from coeval.analyze.main import run_analyze
        run = _make_ees_folder(tmp_path)
        out_dir = tmp_path / 'all_reports'
        code = run_analyze(
            run_path=str(run),
            out_path=str(out_dir),
            subcommand='all',
            agreement_threshold=0.0,
            partial_ok=True,
        )
        assert code == 0
        assert out_dir.exists()
        # Complete report Excel should be created
        assert (out_dir / 'complete_report.xlsx').exists()
        # At least one HTML subfolder should exist
        html_folders = [d for d in out_dir.iterdir() if d.is_dir()]
        assert len(html_folders) >= 1
